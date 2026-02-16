from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from typing import Any

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None

from organizations.models import Region, Club
from registry.models import Athlete, Horse


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _key(v: Any) -> str:
    """
    Normalize header keys (works with Greek + English).
    Removes spaces, /, ., etc and uppercases.
    """
    s = _norm(v).upper()
    return "".join(ch for ch in s if ch.isalnum())


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = _norm(v)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


ATHLETE_ALIASES = {
    # required
    "ΑΜ": "am",
    "AM": "am",
    "ΕΠΩΝΥΜΟ": "last_name",
    "LASTNAME": "last_name",
    "SURNAME": "last_name",
    "ΟΝΟΜΑ": "first_name",
    "FIRSTNAME": "first_name",
    "NAME": "first_name",
    # optional
    "ΠΑΤΡΩΝΥΜΟ": "father_name",
    "FATHERNAME": "father_name",
    "ΗΜΕΡΝΙΑΓΕΝΝΗΣΗΣ": "birth_date",
    "ΗΜΕΡΝΙΑΓΕΝΝΗΣΕΩΣ": "birth_date",
    "BIRTHDATE": "birth_date",
    "ΕΓΓΡΑΦΗΣ": "registration_date",
    "ΗΜΝΙΑΕΓΓΡΑΦΗΣ": "registration_date",
    "REGISTRATIONDATE": "registration_date",
    "ΥΠΗΚΟΟΤΗΤΑ": "nationality",
    "NATIONALITY": "nationality",
    "ΟΜΙΛΟΣ": "club_code",
    "CLUB": "club_code",
}

HORSE_ALIASES = {
    # required
    "ΑΜ": "am",
    "AM": "am",
    "ΙΠΠΟΣ": "name",
    "HORSE": "name",
    "NAME": "name",
    # optional
    "ΔΙΑΒΑΤΗΡΙΟ": "passport_number",
    "PASSPORT": "passport_number",
    "ΗΜΕΡΝΙΑΓΕΝΝΗΣΕΩΣ": "birth_date",
    "ΗΜΕΡΝΙΑΓΕΝΝΗΣΗΣ": "birth_date",
    "BIRTHDATE": "birth_date",
}


def _build_colmap(ws, aliases: dict[str, str], max_scan_rows: int = 40) -> tuple[int, dict[str, int]]:
    """
    Finds header row and returns:
    (header_row_index, internal_field -> column_index)
    """
    max_col = ws.max_column

    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        internal_to_col: dict[str, int] = {}
        for c in range(1, max_col + 1):
            k = _key(ws.cell(row=r, column=c).value)
            if not k:
                continue
            if k in aliases:
                internal_to_col[aliases[k]] = c

        # decide if this row is the header row
        if {"am", "last_name", "first_name"}.issubset(set(internal_to_col.keys())):
            return r, internal_to_col
        if {"am", "name"}.issubset(set(internal_to_col.keys())):
            return r, internal_to_col

    return -1, {}


def _model_field_names(Model) -> set[str]:
    return {f.name for f in Model._meta.fields}


class Command(BaseCommand):
    help = "Import athletes.xlsx + horses.xlsx (demo sync) from settings paths."

    def add_arguments(self, parser):
        parser.add_argument("--athletes-only", action="store_true", help="Import only athletes.xlsx")
        parser.add_argument("--horses-only", action="store_true", help="Import only horses.xlsx")

    @transaction.atomic
    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl is not installed. Run: py -m pip install openpyxl")

        athletes_only = bool(options.get("athletes_only"))
        horses_only = bool(options.get("horses_only"))

        athletes_path = Path(getattr(settings, "ATHLETES_XLSX_PATH", ""))
        horses_path = Path(getattr(settings, "HORSES_XLSX_PATH", ""))

        if not athletes_only and not horses_only:
            # import both
            self._import_athletes(athletes_path)
            self._import_horses(horses_path)
        elif athletes_only:
            self._import_athletes(athletes_path)
        else:
            self._import_horses(horses_path)

        self.stdout.write(self.style.SUCCESS("OK - Excel import completed."))

    def _import_athletes(self, path: Path) -> None:
        if not path or not path.exists():
            raise CommandError(f"athletes.xlsx not found: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        header_row, colmap = _build_colmap(ws, ATHLETE_ALIASES)
        if header_row == -1 or not {"am", "last_name", "first_name"}.issubset(colmap.keys()):
            raise CommandError("Δεν βρήκα header row για athletes (χρειάζεται: ΑΜ, ΕΠΩΝΥΜΟ, ΟΝΟΜΑ).")

        # default region/club for demo mode
        default_region, _ = Region.objects.get_or_create(
            name="UNASSIGNED",
            defaults={"is_active": True},
        )

        athlete_fields = _model_field_names(Athlete)

        created = updated = 0
        for r in range(header_row + 1, ws.max_row + 1):
            am = _norm(ws.cell(row=r, column=colmap["am"]).value)
            if not am:
                continue

            last_name = _norm(ws.cell(row=r, column=colmap.get("last_name")).value)
            first_name = _norm(ws.cell(row=r, column=colmap.get("first_name")).value)
            father_name = _norm(ws.cell(row=r, column=colmap.get("father_name")).value) if colmap.get("father_name") else ""
            birth_date = _parse_date(ws.cell(row=r, column=colmap.get("birth_date")).value) if colmap.get("birth_date") else None
            registration_date = _parse_date(ws.cell(row=r, column=colmap.get("registration_date")).value) if colmap.get("registration_date") else None
            nationality = _norm(ws.cell(row=r, column=colmap.get("nationality")).value) if colmap.get("nationality") else ""
            club_code = _norm(ws.cell(row=r, column=colmap.get("club_code")).value) if colmap.get("club_code") else ""

            if club_code:
                club, _ = Club.objects.get_or_create(
                    code=club_code,
                    defaults={"name": club_code, "region": default_region, "is_active": True},
                )
            else:
                club, _ = Club.objects.get_or_create(
                    code="UNASSIGNED",
                    defaults={"name": "UNASSIGNED", "region": default_region, "is_active": True},
                )

            defaults = {
                "first_name": first_name,
                "last_name": last_name,
                "father_name": father_name,
                "birth_date": birth_date,
                "registration_date": registration_date,
                "nationality": nationality,
                "club": club,
                "is_active": True,
            }

            # keep only fields that exist in your current Athlete model
            defaults = {k: v for k, v in defaults.items() if k in athlete_fields}

            obj, was_created = Athlete.objects.update_or_create(
                registry_number=am,
                defaults=defaults,
            )
            created += 1 if was_created else 0
            updated += 0 if was_created else 1

        self.stdout.write(self.style.SUCCESS(f"Athletes import: created={created}, updated={updated}"))

    def _import_horses(self, path: Path) -> None:
        if not path or not path.exists():
            raise CommandError(f"horses.xlsx not found: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        header_row, colmap = _build_colmap(ws, HORSE_ALIASES)
        if header_row == -1 or not {"am", "name"}.issubset(colmap.keys()):
            raise CommandError("Δεν βρήκα header row για horses (χρειάζεται: ΑΜ, ΙΠΠΟΣ).")

        horse_fields = _model_field_names(Horse)

        created = updated = 0
        for r in range(header_row + 1, ws.max_row + 1):
            am = _norm(ws.cell(row=r, column=colmap["am"]).value)
            if not am:
                continue

            name = _norm(ws.cell(row=r, column=colmap.get("name")).value)
            passport_number = _norm(ws.cell(row=r, column=colmap.get("passport_number")).value) if colmap.get("passport_number") else ""
            birth_date = _parse_date(ws.cell(row=r, column=colmap.get("birth_date")).value) if colmap.get("birth_date") else None

            defaults = {
                "name": name,
                "passport_number": passport_number,
                "birth_date": birth_date,
                "is_active": True,
            }
            defaults = {k: v for k, v in defaults.items() if k in horse_fields}

            obj, was_created = Horse.objects.update_or_create(
                registry_number=am,
                defaults=defaults,
            )
            created += 1 if was_created else 0
            updated += 0 if was_created else 1

        self.stdout.write(self.style.SUCCESS(f"Horses import: created={created}, updated={updated}"))
