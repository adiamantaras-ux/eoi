from datetime import datetime
from django.db import transaction
from django.core.management.base import CommandError
from openpyxl import load_workbook

from registry.models import Athlete
from organizations.models import Club, Region


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except Exception:
            pass
    return None


def get_default_region():
    region, _ = Region.objects.get_or_create(
        name="Χωρίς Περιφέρεια",
        defaults={"is_active": True},
    )
    return region


def get_or_create_club(code):
    code = normalize(code)
    if not code:
        return None

    region = get_default_region()

    club, _ = Club.objects.get_or_create(
        code=code,
        defaults={
            "name": code,
            "region": region,
            "is_active": True,
        },
    )
    return club


@transaction.atomic
def import_athletes_from_xlsx(xlsx_path, stdout=None):
    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb.active

    headers = [normalize(c.value) for c in ws[1]]

    try:
        idx_am = headers.index("ΑΜ")
    except ValueError:
        raise CommandError("Δεν βρέθηκε στήλη 'ΑΜ' στο Excel.")

    idx_last = headers.index("ΕΠΩΝΥΜΟ")
    idx_first = headers.index("ΟΝΟΜΑ")
    idx_father = headers.index("ΠΑΤΡΩΝΥΜΟ")
    idx_birth = headers.index("ΗΜΕΡ/ΝΙΑ ΓΕΝΝΗΣΗΣ")
    idx_nat = headers.index("ΥΠΗΚΟΟΤΗΤΑ")
    idx_club = headers.index("ΟΜΙΛΟΣ")

    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        eoi = normalize(row[idx_am])
        if not eoi:
            continue

        club = get_or_create_club(row[idx_club])

        Athlete.objects.update_or_create(
            eoi_registry_number=eoi,
            defaults={
                "last_name": normalize(row[idx_last]),
                "first_name": normalize(row[idx_first]),
                "father_name": normalize(row[idx_father]),
                "birth_date": parse_date(row[idx_birth]),
                "nationality": normalize(row[idx_nat]),
                "club": club,
            },
        )
        count += 1

    if stdout:
        stdout.write(f"OK. Athletes upserted: {count}")
