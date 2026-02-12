from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.db import transaction

from registry.models import Horse


def _clean(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _parse_date(v: Any) -> Optional[datetime.date]:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        try:
            return datetime(v.year, v.month, v.day).date()
        except Exception:
            return None
    s = _clean(v)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _find_header_row(ws, max_scan_rows: int = 25) -> int:
    targets = {"ΑΜ", "AM"}
    for r in range(1, max_scan_rows + 1):
        row = [_clean(c.value).upper() for c in ws[r]]
        if any(t in row for t in targets) and ("ΙΠΠΟΣ" in row or "HORSE" in row or "NAME" in row):
            return r
    return 1


def _col_index(headers_upper, *names):
    for n in names:
        nu = n.upper()
        if nu in headers_upper:
            return headers_upper.index(nu)
    return None


def import_horses_from_xlsx(path: Path, stdout=None) -> int:
    path = Path(path)
    wb = load_workbook(path)
    ws = wb.active

    header_row = _find_header_row(ws)
    headers = [_clean(c.value) for c in ws[header_row]]
    headers_upper = [h.upper() for h in headers]

    i_am = _col_index(headers_upper, "ΑΜ", "AM")
    i_name = _col_index(headers_upper, "ΙΠΠΟΣ", "HORSE", "NAME")
    i_pass = _col_index(headers_upper, "ΔΙΑΒΑΤΗΡΙΟ", "PASSPORT", "PASSPORT_NUMBER")
    i_birth = _col_index(headers_upper, "ΗΜΕΡ/ΝΙΑ ΓΕΝΝΗΣΕΩΣ", "ΗΜΕΡΟΜΗΝΙΑ ΓΕΝΝΗΣΕΩΣ", "BIRTH_DATE", "DOB")

    horse_fields = {f.name for f in Horse._meta.fields}
    key_field = "registry_number" if "registry_number" in horse_fields else None

    if stdout:
        stdout.write(f"Headers row: {header_row} | Key field: {key_field or '(none)'}")

    count = 0

    with transaction.atomic():
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            am = _clean(row[i_am]) if i_am is not None else ""
            name = _clean(row[i_name]) if i_name is not None else ""
            if not am and not name:
                continue

            defaults = {}
            if name and "name" in horse_fields:
                defaults["name"] = name

            if key_field and am:
                obj, created = Horse.objects.update_or_create(**{key_field: am}, defaults=defaults)
            else:
                obj = Horse.objects.create(**defaults)
                created = True

            if key_field and am:
                setattr(obj, key_field, am)

            if i_pass is not None and "passport_number" in horse_fields:
                obj.passport_number = _clean(row[i_pass])

            if i_birth is not None:
                if "birth_date" in horse_fields:
                    obj.birth_date = _parse_date(row[i_birth])

            obj.save()
            count += 1

    return count


class Command(BaseCommand):
    help = "Import horses from horses.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Path to horses.xlsx")

    def handle(self, *args, **options):
        p = Path(options["path"]).expanduser().resolve()
        n = import_horses_from_xlsx(p, stdout=self.stdout)
        self.stdout.write(self.style.SUCCESS(f"OK. Horses upserted: {n}"))
